# -*- coding: utf-8 -*-
"""
Created on Wed Jan  7 16:57:18 2026

@author: Krishna
"""
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import load_workbook
import re
from openpyxl.utils import get_column_letter

def extract_left_padding(style):
    """Extract the left padding value in points from a style string"""
    if not style:
        return 0
    
    match = re.search(r'padding:\s*\d+\w+\s+\d+\w+\s+\d+\w+\s+(\d+)pt', style)
    if match:
        return int(match.group(1))
    
    # Try padding-left
    match = re.search(r'padding-left:\s*(\d+)pt', style)
    if match:
        return int(match.group(1))
    
    return 0

def calculate_indent_level(left_padding, base_padding=10, indent_step=9):
    """
    Calculate indent level based on left padding.
    base_padding: the baseline padding (usually around 10pt)
    indent_step: typical indent increment (often 9pt or 18pt)
    """
    if left_padding <= base_padding:
        return 0
    
    # Calculate how many indent steps beyond base
    extra_padding = left_padding - base_padding
    indent_level = (extra_padding + indent_step // 2) // indent_step  # Round to nearest
    
    return indent_level

def has_background_color(row):
    """Check if any cell in the row has a non-white background-color attribute"""
    cells = row.find_all('td')
    for cell in cells:
        style = cell.get('style', '')
        if 'background-color' in style:
            # Extract the background color value
            import re
            match = re.search(r'background-color:\s*([^;]+)', style)
            if match:
                color = match.group(1).strip().lower()
                # Ignore white backgrounds
                if color not in ['#ffffff', '#fff', 'white', 'rgb(255,255,255)', 'rgba(255,255,255,1)']:
                    return True
    return False

def parse_html(html_content):
    """Parse HTML table and return JSON structure"""
    soup = BeautifulSoup(html_content, 'html.parser')
    
    table = soup.find('table')
    if not table:
        raise ValueError("No table found in HTML")
    
    rows = table.find_all('tr')
    
    # Always handle rowspans from row 1 into row 2 first
    modified_row_2_cells = None
    if len(rows) > 2:
        # Check row at index 1 for rowspans
        row_1 = rows[1]
        row_1_cells = row_1.find_all('td')
        
        # Track which positions in row 2 need rowspan cells inserted
        rowspan_cells = []  # List of (position, cell, colspan) tuples
        current_pos = 0
        
        for cell in row_1_cells:
            colspan = int(cell.get('colspan', '1'))
            rowspan = int(cell.get('rowspan', '1'))
            
            if rowspan > 1:  # This cell spans into row 2
                rowspan_cells.append((current_pos, cell, colspan))
            
            current_pos += colspan
        
        # If there are rowspan cells, modify row 2
        if rowspan_cells:
            row_2 = rows[2]
            row_2_cells = list(row_2.find_all('td'))
            
            # Build new list with rowspan cells inserted at correct positions
            new_row_2_cells = []
            row_2_cell_idx = 0
            
            # Track how many columns we've processed from actual row 2 cells
            actual_cols_used = 0
            
            for pos, rowspan_cell, rowspan_colspan in rowspan_cells:
                # Add actual row 2 cells until we reach the rowspan position
                while actual_cols_used < pos and row_2_cell_idx < len(row_2_cells):
                    actual_cell = row_2_cells[row_2_cell_idx]
                    new_row_2_cells.append(actual_cell)
                    actual_cols_used += int(actual_cell.get('colspan', '1'))
                    row_2_cell_idx += 1
                
                # Insert the rowspan cell
                new_row_2_cells.append(rowspan_cell)
                actual_cols_used += rowspan_colspan
            
            # Add remaining row 2 cells
            while row_2_cell_idx < len(row_2_cells):
                new_row_2_cells.append(row_2_cells[row_2_cell_idx])
                row_2_cell_idx += 1
            
            # Store the modified cells for row 2
            modified_row_2_cells = new_row_2_cells
    
    # Determine which row to use for column spans
    header_row_index = None
    start_processing_index = None
    
    # Find the first row with gray background color
    first_gray_row_index = None
    for i, row in enumerate(rows):
        if has_background_color(row):
            first_gray_row_index = i
            break
    
    if first_gray_row_index is not None:
        # Go back 2 rows from the first gray row
        header_row_index = max(0, first_gray_row_index - 2)
        start_processing_index = max(1, header_row_index - 1)
    else:
        # Fallback: Search for first row with a non-empty value in the first column
        header_row_index = None
        for i, row in enumerate(rows):
            cells = row.find_all('td')
            if cells:
                first_cell_text = cells[0].get_text(strip=True)
                if first_cell_text:
                    header_row_index = i
                    break
        
        if header_row_index is None:
            header_row_index = 1
        
        start_processing_index = header_row_index
    
    # Get header cells - use modified row 2 if that's the header and we modified it
    if header_row_index == 2 and modified_row_2_cells is not None:
        header_cells = modified_row_2_cells
    else:
        header_cells = rows[header_row_index].find_all('td')
    
    # Build column boundaries based on header colspans
    column_spans = []
    current_pos = 0
    for cell in header_cells:
        colspan = int(cell.get('colspan', '1'))
        column_spans.append((current_pos, current_pos + colspan))
        current_pos += colspan
    
    
    table_data = []
    
    for i, row in enumerate(rows[start_processing_index:], start=start_processing_index):
        # Special handling for row 2 if we modified it
        if i == 2 and modified_row_2_cells is not None:
            cells = modified_row_2_cells
        else:
            cells = row.find_all('td')
        
        current_html_pos = 0  # Position in the raw HTML columns
        ind_to_col_span = {}
        # Check if this row should be center_continuous
        center_continuous = False
        if len(cells) == 1:
            cell = cells[0]
            text = cell.get_text(strip=True)
            colspan = int(cell.get('colspan', '1'))
            if text and colspan > 1:
                center_continuous = True
        
        # Expand cells to their full column positions
        expanded_cells = []
        cell_links = []  # Track hyperlinks for each expanded cell
        first_cell_style = cells[0].get('style', '') if cells else ''
        
        for cell in cells:
            text = cell.get_text(strip=True)
            colspan = int(cell.get('colspan', '1'))
            
            # Check if cell contains a hyperlink
            link_tag = cell.find('a')
            link_url = None
            if link_tag and link_tag.get('href'):
                link_url = link_tag.get('href')
                # Handle relative URLs - prepend base URL if needed
                if link_url.startswith('/'):
                    link_url = 'https://www.sec.gov' + link_url
            
            # Determine which column groups this cell spans
            cell_start_pos = current_html_pos
            cell_end_pos = current_html_pos + colspan
           
            start_group = None
            end_group = None
           
            for group_idx, (col_start, col_end) in enumerate(column_spans):
                if col_start <= cell_start_pos < col_end:
                    start_group = group_idx
                if col_start < cell_end_pos <= col_end:
                    end_group = group_idx
                    break
                elif cell_end_pos > col_end:
                    end_group = group_idx
            
            if start_group is not None and end_group is not None:
                num_groups = end_group - start_group + 1
                if num_groups > 1:
                    ind_to_col_span[start_group] = num_groups
           
           # Expand this cell to its full column positions
            expanded_cells.append(text)
            cell_links.append(link_url)
            
            for _ in range(colspan - 1):
                expanded_cells.append('')
                cell_links.append(None)
           
            current_html_pos += colspan
        
        # Group cells according to column_spans
        grouped_data = []
        grouped_links = []
        #index in grouped data -> num cols in excel they span
        
        for start, end in column_spans:
            group_texts = [expanded_cells[j] for j in range(start, end) if j < len(expanded_cells)]
            combined = ' '.join([t for t in group_texts if t])
            grouped_data.append(combined)
            
            # Get the first non-None link in this group
            group_link = None
            
            for j in range(start, end):
                if j < len(cell_links) and cell_links[j]:
                    group_link = cell_links[j]
                    break
            grouped_links.append(group_link)
        
        # Determine indent level from first cell's left padding
        left_padding = extract_left_padding(first_cell_style)
        indent_level = calculate_indent_level(left_padding)
        table_data.append({
            "data": grouped_data,
            "links": grouped_links,  # Add links to the row data
            "indent_level": indent_level,
            "ind_to_col_span": ind_to_col_span,
            "center_continuous": center_continuous,
            "left_padding": left_padding
        })
    
    # Create JSON object
    json_data = {
        "table": table_data,
        "column_count": len(column_spans),
        "row_count": len(table_data),
        "start_color": header_row_index+2
    }
    
    return json_data



def json_to_excel(json_data, output_file='output.xlsx', hyperlink_url=None, alternating_colors=True):
    """
    Convert parsed table JSON data to a formatted Excel file.
    
    This function takes structured table data (from parse_html_table) and creates an Excel 
    spreadsheet with proper formatting
    
    Args:
        json_data (dict): Parsed table data from parse_html_table containing:
            - 'table': List of row objects with 'data', 'links', 'indent_level', 
                      'center_continuous', and 'left_padding'
            - 'column_count': Number of columns in the table
            - 'row_count': Number of rows in the table
        
        output_file (str, optional): Path where the Excel file will be saved. 
                                    Defaults to 'output.xlsx'.
        
        hyperlink_url (str, optional): URL to hyperlink the first cell (A1) to. 
                                      If A1 is empty, it will be populated with 
                                      "Link to Table". Defaults to None.
        
        alternating_colors (bool, optional): Whether to apply alternating gray/white 
                                            row colors starting from row 2. 
                                            Defaults to True.
    
    Returns:
        None: The function saves the Excel file to disk and doesn't return a value.

    """    
    
    # Prepare data for DataFrame
    rows_for_df = []
    numeric_cells = []
    cell_hyperlinks = []
    
    for row_idx, row_obj in enumerate(json_data['table']):
        data = row_obj['data'].copy()
        links = row_obj.get('links', [None] * len(data))
        indent = row_obj['indent_level']
        center_continuous = row_obj.get('center_continuous', False)
        
        # Add indentation to first column (only if not a center_continuous row)
        if indent > 0 and data[0] and not center_continuous:
            data[0] = '  ' * indent + data[0]
        
        # Track hyperlinks for this row
        for col_idx, link in enumerate(links):
            if link:
                cell_hyperlinks.append({
                    'row': row_idx,
                    'col': col_idx,
                    'url': link
                })
        
        # Convert numeric strings to actual numbers and track their positions
        processed_data = []
        for col_idx, cell in enumerate(data):
            if cell:
                # Check for dollar sign or percent
                has_dollar = '$' in cell
                has_percent = '%' in cell
                has_comma = ',' in cell
                is_negative_parens = False  # Track if number is in parentheses
                
                # Remove formatting for detection
                cleaned = cell.replace(',', '').replace('$', '').replace('%', '').strip()
                
                # Check for parentheses format (negative numbers)
                if cleaned.startswith('(') and cleaned.endswith(')'):
                    is_negative_parens = True
                    cleaned = cleaned[1:-1]  # Remove parentheses
                    cleaned = '-' + cleaned  # Add negative sign
                
                cleaned = re.sub(r'\s+', '', cleaned)
                
                try:
                    # Check if it's a number
                    if re.match(r'^-?\d+\.?\d*$', cleaned):
                        # Determine decimal places
                        decimal_places = 0
                        if '.' in cleaned:
                            decimal_places = len(cleaned.split('.')[1])
                        
                        num_value = float(cleaned) if '.' in cleaned else int(cleaned)
                        
                        # For percentages, divide by 100 since Excel multiplies by 100
                        if has_percent:
                            num_value = num_value / 100
                        
                        processed_data.append(num_value)
                        # Track this cell needs number formatting
                        numeric_cells.append({
                            'row': row_idx,
                            'col': col_idx,
                            'has_dollar': has_dollar,
                            'has_percent': has_percent,
                            'has_comma': has_comma,
                            'is_negative_parens': is_negative_parens,
                            'decimal_places': decimal_places
                        })
                    else:
                        processed_data.append(cell)
                except (ValueError, AttributeError):
                    processed_data.append(cell)
            else:
                processed_data.append(cell)
        
        rows_for_df.append(processed_data)
    
    # Convert to DataFrame
    df = pd.DataFrame(rows_for_df)
    
    # Write to Excel
    df.to_excel(output_file, index=False, header=False)
    
    # Load the workbook with openpyxl for formatting
    wb = load_workbook(output_file)
    ws = wb.active
    ws.sheet_view.showGridLines = False
    
    # Define styles
    header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    header_font = Font(bold=True)
    hyperlink_font = Font(bold=True, color="0563C1", underline="single")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    num_columns = json_data['column_count']
    last_col_letter = get_column_letter(num_columns)
    
    # Apply hyperlinks to cells
    for link_info in cell_hyperlinks:
        cell = ws.cell(row=link_info['row'] + 1, column=link_info['col'] + 1)
        cell.hyperlink = link_info['url']
        cell.font = hyperlink_font
    
    # Apply number formatting to numeric cells
    for cell_info in numeric_cells:
        cell = ws.cell(row=cell_info['row'] + 1, column=cell_info['col'] + 1)
        decimal_places = cell_info['decimal_places']
        has_comma = cell_info['has_comma']
        is_negative_parens = cell_info['is_negative_parens']
        
        # Build format string based on decimal places and comma presence
        if has_comma:
            if decimal_places > 0:
                base_format = f'#,##0.{"0" * decimal_places}'
            else:
                base_format = '#,##0'
        else:
            if decimal_places > 0:
                base_format = f'0.{"0" * decimal_places}'
            else:
                base_format = '0'
        
        # Handle negative numbers in parentheses
        if is_negative_parens:
            # Format: positive;(negative)
            if has_comma:
                if decimal_places > 0:
                    base_format = f'#,##0.{"0" * decimal_places};(#,##0.{"0" * decimal_places})'
                else:
                    base_format = '#,##0;(#,##0)'
            else:
                if decimal_places > 0:
                    base_format = f'0.{"0" * decimal_places};(0.{"0" * decimal_places})'
                else:
                    base_format = '0;(0)'
        
        if cell_info['has_dollar']:
            if is_negative_parens:
                # Format: $positive;($negative)
                if has_comma:
                    if decimal_places > 0:
                        cell.number_format = f'$#,##0.{"0" * decimal_places};($#,##0.{"0" * decimal_places})'
                    else:
                        cell.number_format = '$#,##0;($#,##0)'
                else:
                    if decimal_places > 0:
                        cell.number_format = f'$0.{"0" * decimal_places};($0.{"0" * decimal_places})'
                    else:
                        cell.number_format = '$0;($0)'
            else:
                cell.number_format = f'${base_format}'
        elif cell_info['has_percent']:
            # For percentages, decimal places in the format string
            if decimal_places > 0:
                cell.number_format = f'0.{"0" * decimal_places}%; (0.{"0" * decimal_places})%'
            else:
                cell.number_format = '0%; (0)%'
        else:
            cell.number_format = base_format
    
    
    
    # Format header row (white with bold text)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add hyperlink to first cell if provided
    if hyperlink_url:
        first_cell = ws['A1']
        if not first_cell.value:
            first_cell.value = "Link to Table"
        first_cell.hyperlink = hyperlink_url
        first_cell.font = hyperlink_font
    
    # Format data rows with alternating colors starting from row 2 (if enabled)
    if alternating_colors:
        start_color_row = json_data.get('start_color', 2)
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            # Determine if this row should be gray or white
            # The start_color_row should be gray, then alternate
            offset = row_idx - start_color_row
            if row_idx < start_color_row:
                # Rows before start_color_row are white
                fill = white_fill
            else:
                fill = gray_fill if offset % 2 == 0 else white_fill
            
            for cell in row:
                cell.fill = fill
                if not cell.alignment or (not cell.alignment.horizontal and not cell.alignment.indent):
                    cell.alignment = Alignment(vertical='center')
    else:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for cell in row:
                if not cell.alignment or (not cell.alignment.horizontal and not cell.alignment.indent):
                    cell.alignment = Alignment(vertical='center')
    
    # Apply formatting based on row properties
    for i, row_obj in enumerate(json_data['table'], start=1):
        center_continuous = row_obj.get('center_continuous', False)
        indent = row_obj['indent_level']
        ind_to_col_span = row_obj.get('ind_to_col_span', {})
        
        
        if center_continuous:
            for cell in ws[f'A{i}:{last_col_letter}{i}'][0]:
                cell.alignment = Alignment(horizontal='centerContinuous', vertical='center')
        elif indent > 0:
            ws.cell(row=i, column=1).alignment = Alignment(indent=indent, vertical='center')
        
        for col_idx, num_groups in ind_to_col_span.items():
            start_col_letter = get_column_letter(col_idx + 1)
            end_col_letter = get_column_letter(col_idx + num_groups)
            # Apply centerContinuous to all cells in the range
            for cell in ws[f'{start_col_letter}{i}:{end_col_letter}{i}'][0]:
                cell.alignment = Alignment(horizontal='centerContinuous', vertical='center')
    
    # Auto-adjust column widths
    for column in tuple(ws.columns):
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        
        adjusted_width = min(max_length + 2, 70)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_file)
    wb.close()